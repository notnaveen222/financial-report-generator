# main.py
from openpyxl import Workbook
from sheet_utils.ss_a import build_ssa_sheet
from sheet_utils.ss_b import build_ssb_sheet
from sheet_utils.ss_c import build_ssc_sheet
from sheet_utils.ss_c1 import build_ssc1_sheet
from sheet_utils.ss_c2 import build_ssc2_sheet
from sheet_utils.ss_d import build_ssd_sheet
from sheet_utils.ss_d1 import build_ssd1_sheet
from sheet_utils.styles import (
    headingGreenfill,
    white_bold_font,
    bold_font,
    center_align,
    black_border
)
from openpyxl.styles import Alignment



def generate_excel(payload: dict) -> str:
    
    #splitting data
    landData = payload["landData"]
    buildingData = payload["buildingData"]
    equipmentData = payload["equipmentData"]
    furnitureData=payload["furnitureData"]
    electricData=payload["electricData"]
    otherAssetsData=payload["otherAssetsData"]
    preOperativeExpenseData=payload["preOperativeExpenseData"]
    termLoanPercentage = payload["termLoanPercentage"]
    constructionPeriodData = payload["constructionPeriodData"]
    #sample data
    # ssc_data = [["Indigeneous",161.83], ["Imported", 327.9], ["Erection & Installation Charges 2%", 9.28]]
    # ssc1_data = [["Furniture & Fixtures",161.83]]
    # ssc2_data = [["Total service line charges", 5.55], ["Other Electrical fitting", 9.28]]
    # ssc2_data2 = [["Office equipments", 10.00]]
    # ssd_data = [["Other Expenses", 12.76], ["Lease Rent",1.43], ["Interest during construction period  (SS-D1)", 24.75], ["Bank Charges @0.50%", 2.02]]

    wb = Workbook()
    pc_sheet = wb.create_sheet("Project Cost")
    pc_sheet.column_dimensions["A"].width = 36

    #green heading for project cost
    maxRow = pc_sheet.max_row
    label = pc_sheet.cell(row=maxRow+1, column=1)
    label.value = "Statement-1"
    label.font = white_bold_font
    label.fill = headingGreenfill
    label.alignment = center_align

    maxRow = pc_sheet.max_row
    pc_sheet.merge_cells(start_row=maxRow, start_column=2, end_row=maxRow, end_column=4)
    label = pc_sheet.cell(row=maxRow, column=2)
    label.value = "Project Cost"
    label.font = white_bold_font
    label.fill = headingGreenfill
    label.alignment = center_align
    #

    #SubHeading for Project Cost
    pc_sheet.append(["", "", "", "Rs Lakhs"])
    pc_sheet.cell(row=pc_sheet.max_row, column=4).font = bold_font
    pc_sheet.cell(row=pc_sheet.max_row, column=4).alignment = center_align
    pc_sheet.append([])


    #building sheets
    landTotal = build_ssa_sheet(wb, landData)
    buildingTotal = build_ssb_sheet(wb, buildingData)
    equipmentTotal = build_ssc_sheet(wb, equipmentData["data"])
    furntireTotal = build_ssc1_sheet(wb, furnitureData["data"])
    electricTotal, otherAssetsTotal = build_ssc2_sheet(wb, electricData["data"], otherAssetsData["data"])

    # 1. Calculate D sheet total (excluding D1)
    preOperativeRows = preOperativeExpenseData["data"]
    preOperativeExpenseTotal_excl_D1 = sum(float(row["cost"]) for row in preOperativeRows)

    # 2. For term loan computation, set pre-operative (D) contribution to 100%
    contributionPercentage = [float(item["percent"]) for item in termLoanPercentage["data"]]
    contributionPercentage[-1] = 100.0  # Force last (pre-operative) to 100%
    totals = [
        landTotal, buildingTotal, equipmentTotal, furntireTotal,
        electricTotal, otherAssetsTotal, preOperativeExpenseTotal_excl_D1
    ]
    termLoan_total = 0.0
    for i, value in enumerate(contributionPercentage):
        termLoan = (1 - (value/100)) * totals[i]
        termLoan_total += termLoan

    # 3. Calculate D1 using term loan total
    interestTotalD1 = build_ssd1_sheet(wb, constructionPeriodData, termLoan_total)

    # 4. Build D sheet, passing interestTotalD1 as the D1 row
    preOperativeExpenseTotal = build_ssd_sheet(wb, preOperativeRows, constructionPeriodData, termLoan_total)

    # 5. Update totals dict for reporting
    totals_dict = {
        "Land": landTotal,
        "Building": buildingTotal,
        "Plant & Equipment": equipmentTotal,
        "Furniture & Fittings": furntireTotal,
        "Electrical Fittings": electricTotal,
        "Other Fixed Assets": otherAssetsTotal,
        "Preoperative Expenses": preOperativeExpenseTotal
    }

    pc_sheet.append(["Land", "SS-A","", landTotal])
    pc_sheet.append(["Building", "SS-B","", buildingTotal])
    pc_sheet.append(["Plant & Equipment", "SS-C","", equipmentTotal])
    pc_sheet.append(["Furniture & Fittings", "SS-C1","", furntireTotal])
    #ssc2_total, ssc3_total = build_ssc2_sheet(wb, electricTotal, otherAssetsTotal)
    pc_sheet.append(["Electrical Fittings", "SS-C2","", electricTotal])
    pc_sheet.append(["Other Fixed Assets", "SS-C3","", otherAssetsTotal])
    pc_sheet.append(["Preoperative expenses", "SS-D","", preOperativeExpenseTotal])

    for row in range(5, pc_sheet.max_row+1):
        pc_sheet.cell(row=row, column=2).font = bold_font
        pc_sheet.cell(row=row, column=2).alignment = center_align



    #term load computation table
    pc_sheet.append([])
    # overallTotal1 = 0.0
    # for i, value in enumerate(totals):
    #     overallTotal1+=float(value)
    # pc_sheet.append(["Total", "", "", overallTotal1])
    # maxRow=pc_sheet.max_row



    #term loan total
    maxRow=pc_sheet.max_row
    pc_sheet.merge_cells(start_row=2, start_column=6, end_row=2, end_column=10)
    label = pc_sheet.cell(row=2, column=6)
    label.fill = headingGreenfill
    label.alignment = center_align
    label.font = white_bold_font
    label.value = "Term Loan Computation"

    pc_sheet.column_dimensions["F"].width = 28
    pc_sheet.column_dimensions["H"].width = 22
    pc_sheet.column_dimensions["J"].width = 10

    pc_sheet.row_dimensions[3].height = 32

    cell = pc_sheet.cell(row=3, column=8)
    cell.alignment = Alignment(
        wrap_text=True,
        horizontal='center',
        vertical='center'
    )
    cwr = 3 #current working row
    cwc = 6 #current working col
    tlcHeadings = ["Land", "Building", "Plant & Equipments", "Furniture & Fittings", "Electrical Fittings", "Other Fixed Assets", "Preoperative Expenses"]

    termLoanHeadings = ["", "Rs Lakhs", "% of Contribution to Margin Money", "", "Term Loan"]
    for i, heading in enumerate(termLoanHeadings):
        cell = pc_sheet.cell(row=cwr, column=cwc, value=heading)
        cell.font = bold_font
        cwc+=1
    cwr+=1
    cwr+=1 #for empty space after heading line
    cwc=6
    oldCwc = 6
    oldCwr = cwr
    for i, value in enumerate(tlcHeadings):
        pc_sheet.cell(row=cwr, column=cwc, value=value)
        cwr+=1
    cwr = oldCwr
    cwc = oldCwc+1
    overallTotal = 0.0
    for i, (key, value) in enumerate(totals_dict.items()):
        cell = pc_sheet.cell(row=cwr, column=cwc, value=value)
        cell.number_format = '0.00'
        overallTotal+=value
        cell.font = bold_font
        cell.alignment = center_align
        cwr+=1
    pc_sheet.append(["Total", "", "", overallTotal])
    maxRow=pc_sheet.max_row
    for i in range(4):
        cell = pc_sheet.cell(row=maxRow, column=i+1)
        cell.fill = headingGreenfill
        cell.font = white_bold_font
    
    for row in pc_sheet.iter_rows(min_row=2, max_row=cwr, min_col=oldCwc, max_col=10):
        for cell in row:
            cell.border = black_border

    
    #reposition overall total & ssd1 sheet


    #contributionPercentage = [100, 25, 25, 25, 25, 25, 100]
    contributionPercentage = [float(item["percent"]) for item in termLoanPercentage["data"]]
    cwc+=1
    cwr=oldCwr
    termLoan_total = 0.0

    for i, value in enumerate(contributionPercentage):
        cell = pc_sheet.cell(row=cwr, column=cwc, value=value / 100)  # e.g. 25%
        termLoan = (1 - (value/100)) * (list(totals_dict.values()))[i]
        termLoan_total+=termLoan
        cell.number_format = '0%'  # or '0.00%' if you want 2 decimal places
        cell.font = bold_font
        cell.alignment = center_align
        termLoan_cell = pc_sheet.cell(row=cwr, column=cwc+2, value=termLoan)
        termLoan_cell.number_format = '0.00'
        cwr+=1
    #pc_sheet.append(["", "Rs Lakhs", "% Contribution to Margin Money", "", "Term Loan"])
    cwc=oldCwc
    termLoan_total_row = ["Total", overallTotal, "", "", termLoan_total]
    for i, value in enumerate(termLoan_total_row):
        cell = pc_sheet.cell(row=cwr, column=cwc, value=value)
        cell.font=white_bold_font
        cell.alignment = center_align
        cell.fill = headingGreenfill
        cwc+=1
    cwc=oldCwc

    #termLoanTotoal
    #build_ssd1_sheet(wb, constructionPeriodData,overallTotal)

    #means of finance Table
    pc_sheet.append([])
    maxRow=pc_sheet.max_row
    pc_sheet.merge_cells(start_row=maxRow+2, start_column=2, end_row=maxRow+2, end_column=4)
    maxRow=maxRow+2
    #means of finance
    MOFHeadings = ["Statement-2", "Means of Finance"]
    for i in range(2):
        cell = pc_sheet.cell(row=maxRow, column=i+1)
        cell.value = MOFHeadings[i]
        cell.fill = headingGreenfill
        cell.alignment = center_align
        cell.font = white_bold_font
    pc_sheet.append(["", "", "", "Rs Lakhs"])
    maxRow=pc_sheet.max_row
    cell = pc_sheet.cell(row=maxRow, column=4)
    cell.font = bold_font
    cell.alignment = Alignment(horizontal='right')
    pc_sheet.append([])
    def lastRowDecimalFormatter(columnNumber): #switch this to a utility py file
        maxRow = pc_sheet.max_row
        cell= pc_sheet.cell(row=maxRow, column=columnNumber)
        cell.number_format = '0.00'
    pc_sheet.append(["Equity", "", "", overallTotal-termLoan_total])
    lastRowDecimalFormatter(4)
    pc_sheet.append(["Debt", "", "", termLoan_total])
    lastRowDecimalFormatter(4)
    pc_sheet.append([])
    pc_sheet.append(["Total", "", "", overallTotal])
    lastRowDecimalFormatter(4)
    maxRow = pc_sheet.max_row
    for i in range(4):
        cell = pc_sheet.cell(row=maxRow, column=i+1)
        cell.fill = headingGreenfill
        cell.font= white_bold_font



    #applying border
    for row in pc_sheet.iter_rows(min_row=2, max_row=pc_sheet.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = black_border
    # for row in pc_sheet.iter_rows(min_row=14, max_row=pc_sheet.max_row, min_col=1, max_col=5):
    #     for cell in row:
    #         cell.border = black_border



    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    d1_sheet = wb['SS-D1']
    wb.remove(d1_sheet)
    d11_sheet = wb['SS-D11']
    d11_sheet.title = "SS-D1"
    output_path = "final_report.xlsx"
    wb.save(output_path)
    return output_path

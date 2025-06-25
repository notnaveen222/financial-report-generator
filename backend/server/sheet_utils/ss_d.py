from openpyxl.utils import get_column_letter
from .styles import headingGreenfill, white_font, bold_font, white_bold_font, black_border, center_align
from .ss_d1 import build_ssd1_sheet

def build_ssd_sheet(wb, data, constructionPeriodData, termLoanTotalForD1):
    ssd_sheet = wb.create_sheet("SS-D")
    ssd_sheet.column_dimensions["A"].width = 18
    ssd_sheet.column_dimensions["C"].width = 18

    maxRow = ssd_sheet.max_row
    cell_label = ssd_sheet.cell(row=maxRow+1, column=1)
    cell_label.value = "Sub-Statement D"
    cell_label.font = white_bold_font
    cell_label.fill = headingGreenfill
    ssd_sheet.merge_cells(start_row=maxRow+1, start_column=2, end_row=maxRow+1, end_column=3)
    cell_label = ssd_sheet.cell(row=maxRow+1, column=2)
    cell_label.value = "Pre-Operative Expenses"
    cell_label.font= white_bold_font
    cell_label.fill = headingGreenfill
    cell_label.alignment = center_align

    headings = ["SI.No", "Particulars", "Rs Lakhs/Nos"]
    maxRow=ssd_sheet.max_row
    for col in range(1,4):
        cell = ssd_sheet.cell(row=maxRow+1, column=col)
        cell.value = headings[col-1]
        cell.font = bold_font
        cell.alignment = center_align
    overallTotal = 0.0
    ssd_sheet.append([])

    # 1. Add all D rows except D1
    for i, row in enumerate(data):
        particulars = row["particulars"]
        amount = float(row["cost"])
        overallTotal += amount
        row_data = [i + 1, particulars, amount]
        ssd_sheet.append(row_data)

        row_index = ssd_sheet.max_row
        for j, cell in enumerate(ssd_sheet[row_index], start=1):
            cell.alignment = center_align
            if j == 3:  # Only the amount column
                cell.number_format = "0.00"

    # 2. Calculate D1 and add as a row
    interestTotalD1 = build_ssd1_sheet(wb, constructionPeriodData, termLoanTotalForD1)
    overallTotal += interestTotalD1
    d1_row = [len(data) + 1, "Interest during construction period (SS-D1)", interestTotalD1]
    ssd_sheet.append(d1_row)
    row_index = ssd_sheet.max_row
    for j, cell in enumerate(ssd_sheet[row_index], start=1):
        cell.alignment = center_align
        if j == 3:
            cell.number_format = "0.00"

    # 3. Total row
    totalRow = ["", "Total", overallTotal]
    lastRow = ssd_sheet.max_row
    for col in range(1,4):
        cell = ssd_sheet.cell(row=lastRow+1, column=col)
        cell.value = totalRow[col-1]
        cell.fill = headingGreenfill
        cell.font = white_bold_font
        cell.alignment = center_align

    # Formatting
    max_length = 0
    col_index = 2
    col_letter = get_column_letter(col_index)
    for row in ssd_sheet.iter_rows(min_row=3, max_row=ssd_sheet.max_row, min_col=col_index, max_col=col_index):
        for cell in row:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
    ssd_sheet.column_dimensions[col_letter].width = max_length + 4
    for row in ssd_sheet.iter_rows(min_row=2, max_row=ssd_sheet.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = black_border

    return round(overallTotal, 2)
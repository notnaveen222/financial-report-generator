from openpyxl.utils import get_column_letter
from .styles import headingGreenfill, white_font, bold_font, white_bold_font, black_border, center_align

def build_ssb_sheet(wb, data):
    table_data = data["data"]
    units_data = data["units"]
    ssb_sheet = wb.create_sheet("SS-B")

    # Set column widths
    for col in range(1, 8):
        col_letter = get_column_letter(col)
        ssb_sheet.column_dimensions[col_letter].width = 18

    # Header styling
    for col in range(1, 6):
        cell = ssb_sheet.cell(row=2, column=col)
        cell.fill = headingGreenfill
        cell.font = white_font

    ssb_sheet['A2'] = "Sub-Statement B"
    ssb_sheet['A2'].font = white_bold_font
    ssb_sheet.merge_cells('B2:E2')
    ssb_sheet['B2'] = "Cost of Building"
    ssb_sheet['B2'].font = white_bold_font
    ssb_sheet['B2'].alignment = center_align

    # Table headings
    headings = ["SI.No", "Particulars", f"{units_data["area"]}", f"Rs Lakhs/ {units_data["area"]}", "Rs Lakhs"]
    for col in range(1, 6):
        cell = ssb_sheet.cell(row=3, column=col)
        cell.value = headings[col - 1]
        cell.font = bold_font
        cell.alignment = center_align

    ssb_sheet.append([])

    # Table rows
    overallTotal = 0.0
    #records = data.data
    for i, row in enumerate(table_data):
        name = row["particulars"]
        sqft = float(row["area"])     # treated like area
        cost_per_sqft = float(row["cost"])
        total = sqft * cost_per_sqft
        overallTotal += total
        row_data = [i + 1, name, sqft, cost_per_sqft, total]
        ssb_sheet.append(row_data)

        row_index = ssb_sheet.max_row
        for j, cell in enumerate(ssb_sheet[row_index], start=1):
            cell.alignment = center_align
            if j in [3, 4, 5]:
                cell.number_format = "0.00"

    # Total row
    totalRow = ["", "Total", "", "", overallTotal]
    lastRow = ssb_sheet.max_row + 1
    for col in range(1, 6):
        cell = ssb_sheet.cell(row=lastRow, column=col)
        cell.value = totalRow[col - 1]
        cell.fill = headingGreenfill
        cell.font = white_bold_font
        cell.alignment = center_align

    # Apply borders
    for row in ssb_sheet.iter_rows(min_row=2, max_row=ssb_sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = black_border

    return round(overallTotal, 2)

from openpyxl.utils import get_column_letter
from .styles import headingGreenfill, white_font, bold_font, white_bold_font, black_border, center_align

def build_ssc1_sheet(wb, data):
    ssc1_sheet = wb.create_sheet("SS-C1")
    ssc1_sheet.column_dimensions["A"].width = 18
    ssc1_sheet.column_dimensions["C"].width = 18
    maxRow = ssc1_sheet.max_row
    cell_label = ssc1_sheet.cell(row=maxRow+1, column=1)
    cell_label.value = "Sub-Statement C1"
    cell_label.font = white_bold_font
    cell_label.fill = headingGreenfill
    ssc1_sheet.merge_cells(start_row=maxRow+1, start_column=2, end_row=maxRow+1, end_column=3)
    cell_label = ssc1_sheet.cell(row=maxRow+1, column=2)
    cell_label.value = "Furniture & Fittings"
    cell_label.font= white_bold_font
    cell_label.fill = headingGreenfill
    cell_label.alignment = center_align
    headings = ["SI.No", "Particulars", "Rs Lakhs/Nos"]
    maxRow=ssc1_sheet.max_row
    for col in range(1,4):
        cell = ssc1_sheet.cell(row=maxRow+1, column=col)
        cell.value = headings[col-1]
        cell.font = bold_font
        cell.alignment = center_align
    ssc1_sheet.append([])
    overallTotal = 0.0
    for i, row in enumerate(data):
        particulars = row["particulars"]
        amount = float(row["cost"])
        overallTotal += amount
        row_data = [i + 1, particulars, amount]
        ssc1_sheet.append(row_data)

        row_index = ssc1_sheet.max_row
        for j, cell in enumerate(ssc1_sheet[row_index], start=1):
            cell.alignment = center_align
            if j == 3:  # Only the amount column
                cell.number_format = "0.00"
    totalRow = ["", "Total", overallTotal]
    lastRow = ssc1_sheet.max_row
    for col in range(1,4):
        cell = ssc1_sheet.cell(row=lastRow+1, column=col)
        cell.value = totalRow[col-1]
        cell.fill = headingGreenfill
        cell.font = white_bold_font
        cell.alignment = center_align
    max_length = 0
    col_index = 2
    col_letter = get_column_letter(col_index)

    for row in ssc1_sheet.iter_rows(min_row=3, max_row=ssc1_sheet.max_row, min_col=col_index, max_col=col_index):
        for cell in row:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length

    ssc1_sheet.column_dimensions[col_letter].width = max_length + 4
    for row in ssc1_sheet.iter_rows(min_row=2, max_row=ssc1_sheet.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = black_border
    
    return overallTotal
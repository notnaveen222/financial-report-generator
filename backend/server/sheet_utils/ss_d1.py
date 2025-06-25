from openpyxl.utils import get_column_letter
from .styles import headingGreenfill, white_font, bold_font, white_bold_font, black_border, center_align
from openpyxl.styles import Alignment
from datetime import datetime

def build_ssd1_sheet(wb, data, totalAmount):
    interest = float(data["interest"])
    cod_date = datetime.strptime(data["cod"], "%Y-%m-%d")
    withdrawnPercentage = data["withdrawnPercentage"]

    ssd1_sheet = wb.create_sheet("SS-D1")
    ssd1_sheet.column_dimensions["A"].width = 18
    ssd1_sheet.column_dimensions["B"].width = 14
    ssd1_sheet.column_dimensions["C"].width = 12
    ssd1_sheet.column_dimensions["D"].width = 12
    ssd1_sheet.column_dimensions["E"].width = 12

    # Header setup
    row = ssd1_sheet.max_row + 1
    ssd1_sheet.cell(row=row, column=1, value="Sub-Statement D1").font = white_bold_font
    ssd1_sheet.cell(row=row, column=1).fill = headingGreenfill
    ssd1_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ssd1_sheet.cell(row=row, column=2, value="Interest During Construction Period").font = white_bold_font
    ssd1_sheet.cell(row=row, column=2).fill = headingGreenfill
    ssd1_sheet.cell(row=row, column=2).alignment = center_align

    # Table headings
    row += 1
    ssd1_sheet.row_dimensions[row].height = 32
    headings = ["Month", "% of Loan withdrawn during construction period", "Loan amount withdrawn", "Cumulative Amount Outstanding", "Interest"]
    for i, heading in enumerate(headings):
        cell = ssd1_sheet.cell(row=row, column=i + 1, value=heading)
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        cell.font = bold_font

    # Units row
    row += 1
    ssd1_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ssd1_sheet.cell(row=row, column=2, value="Rs Lakhs").alignment = center_align
    ssd1_sheet.cell(row=row, column=2).font = bold_font

    # Interest value row
    row += 1
    ssd1_sheet.append(["", totalAmount, "", "", interest/100])
    interest_cell = ssd1_sheet.cell(row=ssd1_sheet.max_row, column=5)
    interest_cell.number_format = '0.00%'
    interest_cell.alignment = center_align
    interest_cell.font = bold_font



    # Start data rows
    row = ssd1_sheet.max_row + 1
    cumulative = 0.0
    interest_total = 0.0

    for year, months in withdrawnPercentage.items():
        # Add financial year label
        ssd1_sheet.cell(row=row, column=1, value=year).font = bold_font
        row += 1

        for month, percent_str in months.items():
            percent = float(percent_str)

            # Skip months after COD
            try:
                month_date = datetime.strptime(f"{month} {year.split('-')[0]}", "%B %Y")
                if month_date > cod_date:
                    continue
            except:
                pass  # In case month parsing fails, just include it anyway

            withdrawn_amt = round((percent / 100) * totalAmount, 2)
            cumulative += withdrawn_amt
            monthly_interest = round((cumulative * interest) / (12 * 100), 2)
            interest_total += monthly_interest

            ssd1_sheet.cell(row=row, column=1, value=month.upper())
            ssd1_sheet.cell(row=row, column=2, value=percent / 100).number_format = '0.00%'
            ssd1_sheet.cell(row=row, column=3, value=withdrawn_amt)
            ssd1_sheet.cell(row=row, column=4, value=round(cumulative, 2))
            ssd1_sheet.cell(row=row, column=5, value=monthly_interest)
            row += 1

    # Final Total row
    ssd1_sheet.cell(row=row, column=1, value="Total").font = white_bold_font
    ssd1_sheet.cell(row=row, column=2, value=1.0).number_format = '0.00%'
    ssd1_sheet.cell(row=row, column=3, value=totalAmount)
    ssd1_sheet.cell(row=row, column=5, value=round(interest_total, 2))

    for col in range(1, 6):
        cell = ssd1_sheet.cell(row=row, column=col)
        cell.font = white_bold_font
        cell.fill = headingGreenfill
        cell.alignment = center_align

    # Border for all cells
    for row_cells in ssd1_sheet.iter_rows(min_row=2, max_row=ssd1_sheet.max_row, min_col=1, max_col=5):
        for cell in row_cells:
            cell.border = black_border
    return interest_total
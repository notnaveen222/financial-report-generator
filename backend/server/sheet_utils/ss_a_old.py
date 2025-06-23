from openpyxl.utils import get_column_letter
from .styles import headingGreenfill, white_font, bold_font, white_bold_font, black_border, center_align

def build_ssa_sheet(wb, data):
    ssa_sheet = wb.create_sheet("SS-A")

    # Column width
    for col in range(1, 8):
        col_letter = get_column_letter(col)
        ssa_sheet.column_dimensions[col_letter].width = 18

    # Header styling
    for col in range(1, 6): 
        cell = ssa_sheet.cell(row=2, column=col)
        cell.fill = headingGreenfill
        cell.font = white_font

    ssa_sheet['A2'] = "Sub-Statement A"
    ssa_sheet['A2'].font = white_bold_font
    ssa_sheet.merge_cells('B2:E2')
    ssa_sheet['B2'] = "Land Cost"
    ssa_sheet['B2'].font = white_bold_font
    ssa_sheet['B2'].alignment = center_align

    # Table headings
    headings = ["SI.No", "Particulars", "Acre", "Rs Lakhs/ Acre", "Rs Lakhs"]
    for col in range(1, 6):
        cell = ssa_sheet.cell(row=3, column=col)
        cell.value = headings[col-1]
        cell.font = bold_font
        cell.alignment = center_align

    ssa_sheet.append([])

    # Table rows
    overallTotal = 0.0
    for i, (key, value) in enumerate(data.items()):
        name, acre, rs = value
        total = acre * rs
        overallTotal += total
        row_data = [i + 1, name, acre, rs, total]
        ssa_sheet.append(row_data)

        row_index = ssa_sheet.max_row
        for j, cell in enumerate(ssa_sheet[row_index], start=1):
            cell.alignment = center_align
            if j in [3, 4, 5]:
                cell.number_format = "0.00"

    # Total row
    totalRow = ["", "Total", "", "", overallTotal]
    tempLastRow = ssa_sheet.max_row + 1
    for col in range(1, 6):
        cell = ssa_sheet.cell(row=tempLastRow, column=col)
        cell.value = totalRow[col - 1]
        cell.fill = headingGreenfill
        cell.font = white_bold_font
        cell.alignment = center_align

    # Apply borders
    for row in ssa_sheet.iter_rows(min_row=2, max_row=ssa_sheet.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = black_border

    return overallTotal
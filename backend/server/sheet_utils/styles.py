from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Fills
headingGreenfill = PatternFill(start_color="4f6228", end_color="4f6228", fill_type="solid")
lightGreenfill = PatternFill(start_color="d9ead3", end_color="d9ead3", fill_type="solid")

# Fonts
white_font = Font(color="FFFFFF")
bold_font = Font(bold=True)
white_bold_font = Font(color="FFFFFF", bold=True)

# Alignment
center_align = Alignment(horizontal="center", vertical="center")

# Border
black_border = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)
